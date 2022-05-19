# main script to import and build processed fly dirs

# pyright: reportMissingImports=false, reportMissingModuleSource=false

import os
import json
import numpy as np
import pandas as pd
from shutil import copyfile, rmtree
from xml.etree import ElementTree as ET
from lxml import etree, objectify
from openpyxl import load_workbook, Workbook
import logging
import nibabel as nib
import datetime

from brainsss2.logging_utils import setup_logging, remove_existing_file_handlers # noqa
from brainsss2.argparse_utils import get_base_parser, add_builder_arguments # noqa
from brainsss2.preprocess_utils import dict_to_args_list # noqa
from brainsss2.utils import get_resolution, load_timestamps, sort_nicely # noqa


def parse_args(input, allow_unknown=True):
    parser = get_base_parser('flybuilder')

    parser = add_builder_arguments(parser)

    parser.add_argument(
        "-b", "--basedir",
        type=str,
        help="base directory for fly data",
        required=True)
    parser.add_argument(
        "--func_dirs",
        type=str,
        nargs='+',
        help="specific func dirs to process"
    )
    parser.add_argument('--outstem', type=str, help='output stem')

    if allow_unknown:
        args, unknown = parser.parse_known_args()
        if unknown is not None:
            print(f'skipping unknown arguments:{unknown}')
    else:
        args = parser.parse_args()
    return args


def setup_dirs(args):
    if args.import_dir is None:
        args.import_dir = os.path.join(args.basedir, "imports")
    setattr(args, 'import_path', os.path.join(args.import_dir, args.import_date))
    assert os.path.exists(args.import_path), f"Import path does not exist: {args.import_path}"

    if args.fictrac_import_dir is None:
        args.fictrac_import_dir = os.path.join(args.basedir, "fictrac")
    assert os.path.exists(
        args.fictrac_import_dir
    ), f"fictrac import dir {args.fictrac_import_dir} does not exist"

    if args.target_dir is None:
        args.target_dir = os.path.join(args.basedir, "processed")
    if not os.path.exists(args.target_dir):
        os.mkdir(args.target_dir)
    return(args)


def build_fly(args):

    args = setup_dirs(args)
    print(logging.getLogger())
    print(logging.getLogger().handlers)

    # Assume this folder contains fly1 etc
    # Each area will have a T and a Z
    # Avoid grabbing other weird xml files, reference folder etc.
    # Need to move into fly_X folder that reflects it's date

    # get fly folders in flagged directory and sort to ensure correct fly order
    logging.info(f"Building flies from {args.import_path}")
    likely_fly_folders = os.listdir(args.import_path)
    sort_nicely(likely_fly_folders)
    likely_fly_folders = [i for i in likely_fly_folders if "fly" in i]
    logging.info(f"Found fly folders: {likely_fly_folders}")
    if args.fly_dirs is not None:
        for fly_dir in args.fly_dirs:
            if fly_dir not in likely_fly_folders:
                raise FileNotFoundError(f"specified fly dir {fly_dir} not found in likely fly folders")
        likely_fly_folders = args.fly_dirs
        logging.info(f"Continuing with only{str(likely_fly_folders)}")

    for likely_fly_folder in likely_fly_folders:
        if "fly" in likely_fly_folder:

            setattr(args, 'source_dir', os.path.join(args.import_path, likely_fly_folder))
            new_fly_number = get_new_fly_number(args)

            # Define source fly directory - use args.source_dir instead
            # source_fly = os.path.join(args.import_path, likely_fly_folder)

            # Define destination fly directory
            # fly_time = get_fly_time(source_fly)
            new_fly_name = f"fly_{str(new_fly_number)}"

            setattr(
                args,
                'destination_dir',
                os.path.join(args.target_dir, new_fly_name)
            )
            print(f'flydir: {args.destination_dir}')

            overwrite_msg = None
            if os.path.exists(args.destination_dir):
                if not args.overwrite:
                    logging.warning(f"Fly dir {args.destination_dir} already exists, use -o to overwrite")
                    logging.info(f'flydir: {args.destination_dir}')
                    continue

                overwrite_msg = f"Overwriting existing fly dir {args.destination_dir}"
                rmtree(os.path.join(args.destination_dir))

            os.mkdir(args.destination_dir)

            # print(f'\n*Building {likely_fly_folder} as fly number {new_fly_number}*')
            print(
                f"\n{'   Building '+likely_fly_folder+' as fly_'+ str(new_fly_number) + '   '}"
            )

            # put log file into fly directory
            # args = setup_logging(args, logtype='flybuilder',
            #    logdir=os.path.join(args.destination_dir, "logs"))

            # print(f'Using logger: {logging.getLogger()}')
            logging.info(f"Created fly directory:{args.destination_dir}")
            # print to stdout so that preprocess.py can use this output
            if overwrite_msg is not None:
                logging.info(overwrite_msg)

            # Copy fly data
            copy_fly(args)
            logging.info(f'Copied fly data from {args.source_dir} to {args.destination_dir}')

            # Add date to fly.json file
            try:
                add_date_to_fly(args)
            except Exception as e:
                print(e)

            add_fly_to_conversion_db(args)

            # Add json metadata to master dataset
            try:
                add_fly_to_xlsx(args)
            except Exception as e:
                print(e)


def add_fly_to_conversion_db(args):
    conversion_db_file = os.path.join(args.target_dir, "conversion_db.csv")
    if os.path.exists(conversion_db_file):
        df = pd.read_csv(conversion_db_file)
    else:
        df = pd.DataFrame(columns=['import_dir', 'processed_dir', 'conversion_date'])
    flynum = args.destination_dir.split('/')[-1].split('_')[1]

    df = pd.concat((df, pd.DataFrame.from_records([
        {'flynum': flynum,
        'import_dir': args.source_dir,
        'processed_dir': args.destination_dir,
        'collection_date': args.import_date,
        'collection_session': args.source_dir.split('/')[-1],
        'conversion_date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}])))
    df.to_csv(conversion_db_file, index=False)


def get_new_fly_number(args):
    # first check to see if this fly has already been converted
    conversion_db_file = os.path.join(args.target_dir, "conversion_db.csv")
    new_fly_number = None
    if os.path.exists(conversion_db_file):
        df = pd.read_csv(conversion_db_file, index_col=None)
        if args.source_dir in df.import_dir.values:
            newflydir = df.processed_dir[df.import_dir == args.source_dir].values[0]
            logging.info(f"Fly already converted: {args.source_dir} as {newflydir}")
            new_fly_number = df.flynum[df.import_dir == args.source_dir].values[0]
    if new_fly_number is None:
        oldest_fly = 0
        for current_fly_folder in os.listdir(args.target_dir):
            if current_fly_folder.startswith("fly"):
                fly_num = current_fly_folder.split("_")[-1]
                if int(fly_num) > oldest_fly:
                    oldest_fly = int(fly_num)
        new_fly_number = oldest_fly + 1
    return str(new_fly_number).zfill(3)


# TODO: this seems to have some dangerous assumptions
# and uses bare excepts
def add_date_to_fly(args):
    """ get date from xml file and add to fly.json"""

    # Get date
    # Get func folders
    func_folders = [
        os.path.join(args.destination_dir, x)
        for x in os.listdir(args.destination_dir)
        if "func" in x
    ]
    if len(func_folders) > 0:
        sort_nicely(func_folders)
        func_folder = func_folders[0]
        # Get full xml file path
        xml_file = os.path.join(func_folder, "imaging", "functional.xml")
    else:  # Use anatomy folder
        # Get anat folders
        anat_folders = [
            os.path.join(args.destination_dir, x)
            for x in os.listdir(args.destination_dir)
            if "anat" in x
        ]
        sort_nicely(anat_folders)
        anat_folder = anat_folders[0]
        # Get full xml file path
        xml_file = os.path.join(anat_folder, "imaging", "anatomy.xml")

    # Extract datetime
    datetime_str, _, _ = get_datetime_from_xml(xml_file)
    # Get just date
    date = datetime_str.split("-")[0]
    time = datetime_str.split("-")[1]

    # Add to fly.json
    json_file = os.path.join(args.destination_dir, "fly.json")
    with open(json_file, "r+") as f:
        metadata = json.load(f)
        metadata["date"] = str(date)
        metadata["time"] = str(time)
        f.seek(0)
        json.dump(metadata, f, indent=4)
        f.truncate()


def copy_fly(args):

    """ There will be two types of folders in a fly folder.
    1) func_x folder
    2) anat_x folder
    For functional folders, need to copy fictrac and visual as well
    For anatomy folders, only copy folder. There will also be
    3) fly json data """

    source_fly = args.source_dir
    destination_fly = args.destination_dir

    # look at every item in source fly folder
    for item in os.listdir(source_fly):
        logging.info(f"Currently looking at item: {item}")

        # Handle folders
        if os.path.isdir(os.path.join(source_fly, item)):
            # Call this folder source expt folder
            source_expt_folder = os.path.join(source_fly, item)
            # Make the same folder in destination fly folder
            expt_folder = os.path.join(destination_fly, item)
            os.mkdir(expt_folder)
            logging.info("Created directory: {}".format(expt_folder))

            # Is this folder an anatomy or functional folder?
            if "anat" in item:
                # If anatomy folder, just copy everything
                # Make imaging folder and copy
                imaging_destination = os.path.join(expt_folder, "imaging")
                if not os.path.exists(imaging_destination):
                    os.mkdir(imaging_destination)
                copy_bruker_data(source_expt_folder, imaging_destination, "anat", args)

            elif "func" in item:
                if args.func_dirs is not None and item.split('/')[-1] not in args.func_dirs:
                    logging.info(f"Skipping {item} - not in args.func_dirs")
                    continue
                # Make imaging folder and copy
                imaging_destination = os.path.join(expt_folder, "imaging")
                if not os.path.exists(imaging_destination):
                    os.mkdir(imaging_destination)
                copy_bruker_data(source_expt_folder, imaging_destination, "func", args)
                # Copt fictrac data based on timestamps
                copy_fictrac(expt_folder, args)

            else:
                logging.warning("Invalid directory in fly folder (skipping): {}".format(item))

        # Copy fly.json file
        else:
            if item == "fly.json":
                source_path = os.path.join(source_fly, item)
                target_path = os.path.join(destination_fly, item)
                logging.info('Copying from {} to {}'.format(source_path, target_path))
                copyfile(source_path, target_path)
            else:
                logging.warning("Invalid file in fly folder (skipping): {}".format(item))


def copy_bruker_data(source, destination, folder_type, args):
    # Do not update destination - download all files into that destination
    for item in os.listdir(source):
        # Create full path to item
        source_item = os.path.join(source, item)

        # Check if item is a directory
        if os.path.isdir(source_item):
            # Do not update destination - download all files into that destination
            copy_bruker_data(source_item, destination, folder_type, args)

        # If the item is a file
        else:
            # Change file names and filter various files
            # Don't copy these files
            if "SingleImage" in item:
                continue
            # Rename functional file to functional_channel_x.nii
            if ".nii" in item and folder_type == "func":
                # '_' is from channel numbers my tiff to nii adds
                item = "functional_" + item.split("_")[1] + "_" + item.split("_")[2]
                target_item = os.path.join(destination, item)
                copy_nifti_file(source_item, target_item)
                continue

            # Rename anatomy file to anatomy_channel_x.nii
            if ".nii" in item and folder_type == "anat":
                item = "anatomy_" + item.split("_")[1] + "_" + item.split("_")[2]
                target_item = os.path.join(destination, item)
                if item.find('channel_1') > -1:
                    copy_nifti_file(source_item, target_item)
                else:
                    logging.info('skipping anat channel 2')
                continue

            # Special copy for photodiode since it goes in visual folder
            if ".csv" in item:
                item = "photodiode.csv"
                try:
                    args.visual_import_dir = os.path.join(
                        os.path.split(destination)[0], "visual"
                    )
                    os.mkdir(args.visual_import_dir)
                except (FileExistsError, FileNotFoundError):
                    pass
                target_item = os.path.join(
                    os.path.split(destination)[0], "visual", item
                )
                copyfile(source_item, target_item)
                continue
            # Special copy for visprotocol metadata since it goes in visual folder
            if ".hdf5" in item:
                visual_dir = os.path.join(os.path.split(destination)[0], "visual")
                if not os.path.exists(visual_dir):
                    os.mkdir(visual_dir)
                target_item = os.path.join(visual_dir, item)
                copyfile(source_item, target_item)
                continue
            # Rename to anatomy.xml if appropriate
            if ".xml" in item and folder_type == "anat" and "Voltage" not in item:
                item = "anatomy.xml"
            # Rename to functional.xml if appropriate, copy immediately, then make scan.json
            if ".xml" in item and folder_type == "func" and "Voltage" not in item:
                item = "functional.xml"
                target_item = os.path.join(destination, item)
                copy_file(source_item, target_item)
                # Create json file
                create_imaging_json(target_item, args)
                continue
            if ".xml" in item and "VoltageOutput" in item:
                item = "voltage_output.xml"
            # Special copy for expt.json
            if "expt.json" in item:
                target_item = os.path.join(os.path.split(destination)[0], item)
                copyfile(source_item, target_item)
                continue

            # Actually copy the file
            target_item = os.path.join(destination, item)
            print('copying using copy_file')
            copy_file(source_item, target_item)


def copy_file(source, target):
    logging.info(f'Copy file {source} to {target}')
    copyfile(source, target)


def copy_nifti_file(source, target, stepsize=4):
    """copy nifti file and set header info
    - copy in chunks to prevent memory overload"""
    # from luke:
    # the array is shape (256, 128, 49, 3384).
    # Axis 1: Goes laterally across the brain. Starts on the left side of the brain as if you are the fly.
    # Axis 2: Dorsal-Ventral axis. Starts at dorsal
    # Axis 3: Posterior-Anterior axis. Starts at posterior.
    # Axis 4: time
    logging.info(f'Copying nifti file {source} to {target}')
    xmlfile = source.split('_channel')[0] + '.xml'
    resolution = list(get_resolution(xmlfile))
    timestamps = load_timestamps(os.path.dirname(xmlfile), os.path.basename(xmlfile))
    # resolution.append(np.diff(timestamps[:, 0])[0]/1000)
    resolution.append(1)
    affine = np.diag(np.array(resolution) / 1000)
    orig_img = nib.load(source, mmap='r')

    img = nib.Nifti1Image(orig_img.dataobj, affine=None)
    img.header.set_qform(affine, code=2)
    img.header.set_sform(affine, code=2)
    img.header.set_xyzt_units(xyz='mm', t='sec')
    img.header.set_zooms(list(resolution[:3]) + [np.diff(timestamps[:, 0])[0] / 1000])

    logging.info(f"saving to target: {target}")
    img.to_filename(target)


def copy_fictrac(func_dir, args):
    """find matching fictrac dataset and copy info fictract folder"""

    # fictrac_folder = '/oak/stanford/groups/trc/data/Brezovec/2P_Imaging/imports/fictrac'
    if args.fictrac_import_dir is None:
        args.fictrac_import_dir = os.path.join(args.basedir, "fictrac")
    assert os.path.exists(
        args.fictrac_import_dir
    ), f"fictrac import directory does not exist ({args.fictrac_import_dir})"

    fictrac_destination = os.path.join(func_dir, "fictrac")
    if not os.path.exists(fictrac_destination):
        os.mkdir(fictrac_destination)

    # Find time of experiment based on functional.xml
    true_ymd, true_total_seconds = get_expt_time(os.path.join(func_dir, "imaging"))
    logging.info(f"true_ymd: {true_ymd}; true_total_seconds: {true_total_seconds}")

    # Find .dat file of 1) correct-ish time, 2) correct-ish size
    datetime_correct = None
    for file in os.listdir(args.fictrac_import_dir):

        # must be .dat file
        if ".dat" not in file:
            continue

        # Get datetime from file name
        datetime = file.split("-")[1][:-4]
        test_ymd = datetime.split("_")[0]
        test_time = datetime.split("_")[1]
        test_hour = test_time[:2]
        test_minute = test_time[2:4]
        test_second = test_time[4:6]
        test_total_seconds = (
            int(test_hour) * 60 * 60 + int(test_minute) * 60 + int(test_second)
        )

        # Year/month/day must be exact
        if true_ymd != test_ymd:
            continue
        logging.info(f"Found file from same day: {file}")

        # Time must be within 10min
        time_difference = np.abs(true_total_seconds - test_total_seconds)
        if time_difference > 10 * 60:
            continue
        logging.info("Found fictrac file that matches time.")

        # Must be correct size
        fp = os.path.join(args.fictrac_import_dir, file)
        file_size = os.path.getsize(fp)
        if file_size > 30000000:  # 30MB
            logging.info(f"Found correct .dat file{file}")
            datetime_correct = datetime
            break

    if datetime_correct is None:
        logging.warning(
            f"{'   No fictrac data found --- continuing without fictrac data   '}"
        )
        return

    # Collect all fictrac files with correct datetime
    correct_time_files = [
        file for file in os.listdir(args.fictrac_import_dir) if datetime_correct in file
    ]

    # Now transfer these 4 files to the fly
    for file in correct_time_files:
        target_path = os.path.join(fictrac_destination, file)
        source_path = os.path.join(args.fictrac_import_dir, file)
        copyfile(source_path, target_path)

    # Create empty xml file - Update this later
    root = etree.Element("root")
    fictrac = objectify.Element("fictrac")
    root.append(fictrac)
    objectify.deannotate(root)
    etree.cleanup_namespaces(root)
    tree = etree.ElementTree(fictrac)
    with open(os.path.join(fictrac_destination, "fictrac.xml"), "wb") as file:
        tree.write(file, pretty_print=True)


def create_imaging_json(xml_source_file, args):

    # Get datetime
    try:
        datetime_str, _, _ = get_datetime_from_xml(xml_source_file)
    except FileNotFoundError:
        logging.warning("No xml or cannot read.")
        return

    date = datetime_str.split("-")[0]
    time = datetime_str.split("-")[1]
    source_data = {"date": str(date), "time": str(time)}

    # Get rest of data
    tree = objectify.parse(xml_source_file)
    source = tree.getroot()
    statevalues = source.findall("PVStateShard")[0].findall("PVStateValue")
    for statevalue in statevalues:
        key = statevalue.get("key")
        if key == "micronsPerPixel":
            indices = statevalue.findall("IndexedValue")
            for index in indices:
                axis = index.get("index")
                if axis == "XAxis":
                    source_data["x_voxel_size"] = float(index.get("value"))
                elif axis == "YAxis":
                    source_data["y_voxel_size"] = float(index.get("value"))
                elif axis == "ZAxis":
                    source_data["z_voxel_size"] = float(index.get("value"))
        if key == "laserPower":
            # I think this is the maximum power if set to vary by z depth - WRONG
            indices = statevalue.findall("IndexedValue")
            laser_power_overall = int(float(indices[0].get("value")))
            source_data["laser_power"] = laser_power_overall
        if key == "pmtGain":
            indices = statevalue.findall("IndexedValue")
            for index in indices:
                index_num = index.get("index")
                if index_num == "0":
                    source_data["PMT_red"] = int(float(index.get("value")))
                if index_num == "1":
                    source_data["PMT_green"] = int(float(index.get("value")))
        if key == "pixelsPerLine":
            source_data["x_dim"] = int(float(statevalue.get("value")))
        if key == "linesPerFrame":
            source_data["y_dim"] = int(float(statevalue.get("value")))
    sequence = source.findall("Sequence")[0]
    last_frame = sequence.findall("Frame")[-1]
    source_data["z_dim"] = int(last_frame.get("index"))

    # Save data
    with open(os.path.join(os.path.split(xml_source_file)[0], "scan.json"), "w") as f:
        json.dump(source_data, f, indent=4)


def get_expt_time(directory):
    """ Finds time of experiment based on functional.xml """
    xml_file = os.path.join(directory, "functional.xml")
    _, _, datetime_dict = get_datetime_from_xml(xml_file)
    true_ymd = datetime_dict["year"] + datetime_dict["month"] + datetime_dict["day"]
    true_total_seconds = (
        int(datetime_dict["hour"]) * 60 * 60 +
        int(datetime_dict["minute"]) * 60 +
        int(datetime_dict["second"])
    )

    return true_ymd, true_total_seconds


def get_fly_time(fly_folder):
    # need to read all xml files and pick oldest time
    # find all xml files
    xml_files = []
    xml_files = get_xml_files(fly_folder, xml_files)

    logging.info(f'found xml files: {xml_files}')

    datetimes_str = []
    datetimes_int = []
    for xml_file in xml_files:
        datetime_str, datetime_int, _ = get_datetime_from_xml(xml_file)
        datetimes_str.append(datetime_str)
        datetimes_int.append(datetime_int)

    # Now pick the oldest datetime
    datetimes_int = np.asarray(datetimes_int)
    logging.info(f'Found datetimes: {datetimes_str}')
    index_min = np.argmin(datetimes_int)
    datetime = datetimes_str[index_min]
    logging.info(f'Found oldest datetime: {datetime}')
    return datetime


def get_xml_files(fly_folder, xml_files):
    # Look at items in fly folder
    for item in os.listdir(fly_folder):
        full_path = os.path.join(fly_folder, item)
        if os.path.isdir(full_path):
            xml_files = get_xml_files(full_path, xml_files)
        elif (".xml" in item
              and "_Cycle" not in item
              and "fly.xml" not in item
              and "scan.xml" not in item
              and "expt.xml" not in item):
            xml_files.append(full_path)
            logging.info(f'Found xml file: {full_path}')
    return xml_files


def get_datetime_from_xml(xml_file):
    logging.info(f'Getting datetime from {xml_file}')
    tree = ET.parse(xml_file)
    root = tree.getroot()
    datetime = root.get("date")
    # will look like "4/2/2019 4:16:03 PM" to start

    # Get dates
    date = datetime.split(" ")[0]
    month = date.split("/")[0]
    day = date.split("/")[1]
    year = date.split("/")[2]

    # Get times
    time = datetime.split(" ")[1]
    hour = time.split(":")[0]
    minute = time.split(":")[1]
    second = time.split(":")[2]

    # Convert from 12 to 24 hour time
    am_pm = datetime.split(" ")[-1]
    if am_pm == "AM" and hour == "12":
        hour = str(00)
    elif am_pm != "AM" and (am_pm != "PM" or hour != "12"):
        hour = str(int(hour) + 12)

    # Add zeros if needed
    if len(month) == 1:
        month = "0" + month
    if len(day) == 1:
        day = "0" + day
    if len(hour) == 1:
        hour = "0" + hour

    # Combine
    datetime_str = year + month + day + "-" + hour + minute + second
    datetime_int = int(year + month + day + hour + minute + second)
    datetime_dict = {
        "year": year,
        "month": month,
        "day": day,
        "hour": hour,
        "minute": minute,
        "second": second,
    }

    return datetime_str, datetime_int, datetime_dict


def load_json(file):
    with open(file, "r") as f:
        data = json.load(f)
    return data


def load_xml(file):
    tree = objectify.parse(file)
    return tree.getroot()


# NOTE: this is seems like an inefficient way to do this
# I would save as a standard data frame
# or save as a json into each folder and then create a separate script to generate the
# data frame
def add_fly_to_xlsx(args):
    # try to load excel file
    # TODO: should use a standard format like TSV instead of xlsx
    if args.xlsx_file is None:
        args.xlsx_file = os.path.join(args.basedir, "master_2P.xlsx")
    if os.path.exists(args.xlsx_file):
        wb = load_workbook(filename=args.xlsx_file, read_only=False)
        ws = wb.active
        logging.info(f'Loaded existing xlsx file: {args.xlsx_file}')
    else:
        logging.info(f'Creating new xlsx file: {args.xlsx_file}')
        wb = Workbook()
        ws = wb.active
        title_row = [
            'fly_id',
            'expt_id',
            'date',
            'brain_area',
            "genotype",
            'visual_input',
            'unknown',
            "fly_notes",
            "expt_notes",
            "expt_time",
            "circadian_on",
            "circadian_off",
            "gender",
            "age",
            "temp",
            "laser_power",
            "PMT_green",
            "PMT_red",
            "x_dim",
            "y_dim",
            "z_dim",
            "x_voxel_size",
            "y_voxel_size",
            "z_voxel_size",
        ]
        ws.append(title_row)

    # TRY TO LOAD FLY METADATA
    try:
        fly_file = os.path.join(args.destination_dir, "fly.json")
        fly_data = load_json(fly_file)
    except FileNotFoundError:
        logging.warning("FYI no *fly.json* found; this will not be logged in your excel sheet.")
        fly_data = {
            "circadian_on": None,
            "circadian_off": None,
            "gender": None,
            "age": None,
            "temp": None,
            "notes": None,
            "date": None,
            "genotype": None}

    func_folders = [
        os.path.join(args.destination_dir, x) for x in os.listdir(args.destination_dir) if "func" in x
    ]
    sort_nicely(func_folders)
    for func_folder in func_folders:

        # TRY TO LOAD EXPT METADATA
        try:
            expt_file = os.path.join(func_folder, "expt.json")
            expt_data = load_json(expt_file)
        except FileNotFoundError:
            logging.warning(
                "FYI no *expt.json* found; this will not be logged in your excel sheet."
            )
            expt_data = {"brain_area": None, "notes": None, "time": None}

        # TRY TO LOAD SCAN DATA
        try:
            scan_file = os.path.join(func_folder, "imaging", "scan.json")
            scan_data = load_json(scan_file)
            scan_data["x_voxel_size"] = "{:.1f}".format(scan_data["x_voxel_size"])
            scan_data["y_voxel_size"] = "{:.1f}".format(scan_data["y_voxel_size"])
            scan_data["z_voxel_size"] = "{:.1f}".format(scan_data["z_voxel_size"])
        except FileNotFoundError:
            scan_data = {
                "laser_power": None,
                "PMT_green": None,
                "PMT_red": None,
                "x_dim": None,
                "y_dim": None,
                "z_dim": None,
                "x_voxel_size": None,
                "y_voxel_size": None,
                "z_voxel_size": None}

        visual_file = os.path.join(func_folder, "visual", "visual.json")
        try:
            visual_data = load_json(visual_file)
            visual_input = visual_data[0]["name"] + f" ({len(visual_data)})"
        except FileNotFoundError:
            visual_input = None

        # Get fly_id
        fly_id = args.destination_dir.split("_")[-1]

        # Get expt_id
        expt_id = func_folder.split("_")[-1]

        # Append the new row
        new_row = [
            int(fly_id),
            int(expt_id),
            fly_data["date"],
            expt_data["brain_area"],
            fly_data["genotype"],
            visual_input,
            None,
            fly_data["notes"],
            expt_data["notes"],
            expt_data["time"],
            fly_data["circadian_on"],
            fly_data["circadian_off"],
            fly_data["gender"],
            fly_data["age"],
            fly_data["temp"],
            scan_data["laser_power"],
            scan_data["PMT_green"],
            scan_data["PMT_red"],
            scan_data["x_dim"],
            scan_data["y_dim"],
            scan_data["z_dim"],
            scan_data["x_voxel_size"],
            scan_data["y_voxel_size"],
            scan_data["z_voxel_size"],
        ]

        ws.append(new_row)

    # Save the file
    wb.save(args.xlsx_file)


# for use from outside
def build_fly_from_argdict(argdict):
    arglist = dict_to_args_list(argdict)
    args = parse_args(arglist)
    build_fly(args)
